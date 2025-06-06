# Estilo_excel.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import locale

# Detectar la configuración regional del sistema
region = locale.getdefaultlocale()[0]

# Definir formato de fecha dinámico según la región
if region and region.startswith("es_"):
    formato_fecha = 'DD/MM/YYYY HH:MM:SS'
else:
    formato_fecha = 'YYYY-MM-DD HH:MM:SS'


def aplicar_estilo_encabezado(cell):
    cell.font = Font(color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4F81BD")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))


def crear_excel_con_estilo(registros, ruta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    # Encabezados
    headers = ["Tipo", "Descripción", "Monto", "Categoría", "Fecha"]
    ws.append(headers)

    # Estilos
    header_font = Font(color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Aplicar estilo a los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        aplicar_estilo_encabezado(cell)

    # Insertar registros con formato
    current_row = 2
    for r in registros:
        row_data = [r.tipo, r.descripcion, r.monto, r.categoria, r.fecha]

        # Validación: ignorar registros incompletos
        if not all(row_data):
            continue

        fill_color = PatternFill("solid", fgColor="D9E1F2" if current_row % 2 == 0 else "FFFFFF")

        for col_num, valor in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num, value=valor)
            cell.fill = fill_color
            cell.alignment = alignment
            cell.border = thin_border

            if col_num == 5:  # Columna Fecha
                try:
                    fecha_dt = datetime.strptime(valor, "%Y-%m-%d %H:%M:%S")
                    cell.value = fecha_dt
                    cell.number_format = formato_fecha
                except ValueError:
                    pass

        current_row += 1  # Solo avanzar si la fila fue válida

    # Ajustar anchos de columnas
    column_widths = {
        "A": 12,
        "B": 40,
        "C": 12,
        "D": 20,
        "E": 25,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(ruta)
